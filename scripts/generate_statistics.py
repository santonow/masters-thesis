import json
import os
from collections import defaultdict, Counter
from functools import partial
from itertools import combinations, chain
import sys
from typing import Dict, Callable, Iterator, Union, Any, List, Optional, Tuple, Set
import csv
from math import isnan

import networkx as nx
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from joblib import Parallel, delayed


RANKS = [
    "kingdom",
    "supergroup",
    "division",
    "class",
    "order",
    "family",
    "genus",
    "species",
]

Taxonomy = tuple[str, ...]


def dict_zip(*dicts: dict, default=None) -> Tuple[Any, ...]:
    keys = set()
    for d in dicts:
        keys.update(d)
    for key in keys:
        yield tuple([key] + [d.get(key, default) for d in dicts])


class OpenPyXLWriter:
    def __init__(self, fname: str, sheet_name: Optional[str] = None):
        self.fname = fname
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        if sheet_name is not None:
            self.ws.title = sheet_name
        self.row_count = {self.ws.title: 1}

    def _add_sheet(self, sheet_name: str):
        self.wb.create_sheet(title=sheet_name)
        self.row_count[sheet_name] = 1

    def _switch_sheet(self, sheet_name: str):
        if sheet_name not in self.row_count:
            self._add_sheet(sheet_name)
        self.ws = self.wb[sheet_name]

    def _update_row(self):
        self.row_count[self.ws.title] += 1

    @property
    def _current_pos(self):
        return self.row_count[self.ws.title]

    def _pos_iter(self):
        i = 1
        while True:
            yield f"{get_column_letter(i)}{self._current_pos}"
            i += 1

    def writerow(self, row: List[Any], sheet_name: Optional[str] = None):
        if sheet_name is not None and self.ws.title != sheet_name:
            self._switch_sheet(sheet_name)
        for elem, pos in zip(row, self._pos_iter()):
            try:
                if elem is None:
                    self.ws[pos] = str(elem)
                else:
                    self.ws[pos] = elem
            except ValueError as e:
                self.ws[pos] = str(elem)
        self._update_row()

    def save(self):
        for sheet_name, count in self.row_count.items():
            if count == 1:
                del self.wb[sheet_name]
        self.wb.save(self.fname)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()


def prepare_hub_taxons_stats(
    graphs: List[nx.Graph],
    names: List[str],
    taxonomy: Dict[str, Taxonomy],
    levels: Tuple[int, ...] = (4, 5, 6, 7),
) -> Dict[int, List[List]]:
    results = dict()
    for level in levels:
        new_taxonomy = {otu_id: lineage[:level] for otu_id, lineage in taxonomy.items()}

        def get_tax_ranking(graph):
            tax2ranks = dict()
            tax2degrees = Counter()
            for node, degree in graph.degree():
                if node == "leftover_vector":
                    continue
                tax2degrees[new_taxonomy[node]] += degree
            for i, (tax, degree) in enumerate(
                sorted(tax2degrees.items(), reverse=True, key=lambda x: x[1])
            ):
                tax2ranks[tax] = {"degree": degree, "rank": i + 1}
            return tax2ranks

        tax2degrees = defaultdict(lambda: defaultdict(Counter))
        for tax, *degrees in dict_zip(*[get_tax_ranking(graph) for graph in graphs]):
            for graph_name, data in zip(names, degrees):
                if data is not None:
                    tax2degrees[tax][graph_name] += data

        rows = []
        for node, vals in tax2degrees.items():
            #     if not set(vals) == set()
            row = [node]
            degrees = [val["degree"] for val in vals.values()]
            ranks = [val["rank"] for val in vals.values()]
            for graph_name in names:
                row.append(vals.get(graph_name, dict()).get("degree"))
                row.append(vals.get(graph_name, dict()).get("rank"))
            row.append(np.mean(degrees))
            row.append(np.mean(ranks))
            rows.append(row)

        results[level] = sorted(rows, key=lambda x: x[-1])

    return results


def write_hub_stats(
    handle: OpenPyXLWriter, names: List[str], hub_stats: Dict[int, List[List]]
):
    for level, rows in hub_stats.items():
        sheet_name = f"{RANKS[level - 1]} sorted by degree"
        trimmed_ranks = RANKS[:level]
        header = list(trimmed_ranks)
        for name in names:
            header.append(f"degree ({name})")
            header.append(f"rank ({name})")
        header.append("mean degree")
        header.append("mean rank")
        handle.writerow(header, sheet_name=sheet_name)
        for row in sorted(rows, key=lambda x: x[-1]):
            tax, *vals = row
            tax = list(tax) + ["" for _ in range(len(trimmed_ranks) - len(tax))]
            assert len(tax) == len(trimmed_ranks)
            handle.writerow(list(tax) + vals, sheet_name=sheet_name)


def prepare_pairwise_stats(
    method_graphs: dict[str, nx.Graph],
    handle: OpenPyXLWriter,
    taxonomy: dict[str, tuple[str, ...]],
):
    def common_taxon_edges(c1: Counter, c2: Counter) -> int:
        common = 0
        for k in set(c1) & set(c2):
            common += min([c1[k], c2[k]])
        return common

    edges_OTU = {
        method: set(tuple(sorted([head, tail])) for head, tail in graph.edges())
        for method, graph in method_graphs.items()
    }
    edges_taxon = {
        method: Counter(
            tuple(sorted([taxonomy[head], taxonomy[tail]]))
            for head, tail in graph.edges()
        )
        for method, graph in method_graphs.items()
    }
    edges_genus = dict()
    for method, graph in method_graphs.items():
        c = Counter()
        for head, tail in graph.edges():
            head_taxonomy = taxonomy[head]
            tail_taxonomy = taxonomy[tail]
            if len(head_taxonomy) == 8:
                head_taxonomy = head_taxonomy[:-1]
            if len(tail_taxonomy) == 8:
                tail_taxonomy = tail_taxonomy[:-1]
            c[tuple(sorted([head_taxonomy, tail_taxonomy]))] += 1
        edges_genus[method] = c

    common_OTUs = {
        tuple(sorted([method1, method2])): len(edges_OTU[method1] & edges_OTU[method2])
        for method1, method2 in combinations(method_graphs, 2)
    }
    common_taxons = {
        tuple(sorted([method1, method2])): common_taxon_edges(
            edges_taxon[method1], edges_taxon[method2]
        )
        for method1, method2 in combinations(method_graphs, 2)
    }
    common_genus = {
        tuple(sorted([method1, method2])): common_taxon_edges(
            edges_genus[method1], edges_genus[method2]
        )
        for method1, method2 in combinations(method_graphs, 2)
    }

    methods = sorted(method_graphs)
    handle.writerow([""] + methods, sheet_name="Method comparisons")
    for level, common_edges in [
        ("OTU level", common_OTUs),
        ("taxon level", common_taxons),
        ("genus level", common_genus),
    ]:
        for method in methods:
            row = [f"{method} ({level})"]
            for other_method in methods:
                if method == other_method:
                    row.append("")
                else:
                    row.append(common_edges[tuple(sorted([method, other_method]))])

            handle.writerow(row, sheet_name="Method comparisons")


def read_trophic_groups(fpath, taxonomy):
    trophic_groups = pd.read_excel(fpath)
    taxa_to_trophic_group = dict()

    for row in trophic_groups.to_dict(orient="records"):
        if isinstance(row["Trophic group"], float):
            continue
        else:
            for i in reversed(range(1, 6)):
                if not (
                    isinstance(row[f"Taxa_{i}"], float) and isnan(row[f"Taxa_{i}"])
                ):
                    taxa_to_trophic_group[row[f"Taxa_{i}"].strip()] = (
                        row["Trophic group"].strip().replace("parasitic", "parasite")
                    )
    group_to_taxa = defaultdict(set)
    for taxa, group in taxa_to_trophic_group.items():
        group_to_taxa[group].add(taxa)
    otu_to_group = dict()
    for otu_id, lineage in taxonomy.items():
        for trophic_group, taxons in group_to_taxa.items():
            if set(lineage) & taxons:
                otu_to_group[otu_id] = trophic_group
    return otu_to_group


def read_experimental_interactions(
    interactions_fpath: str,
    taxons_in_otu_table: set[Taxonomy],
    trim_to_genus: bool = False,
) -> dict[Tuple[Taxonomy, Taxonomy], Set[str]]:
    interactions = defaultdict(set)
    with open(interactions_fpath) as handle:
        reader = csv.reader(handle, delimiter="\t")
        next(reader)
        for head, tail, interaction in reader:
            head_lineage = tuple(json.loads(head))
            tail_lineage = tuple(json.loads(tail))
            if trim_to_genus:
                if len(head_lineage) == 8:
                    head_lineage = head_lineage[:-1]
                if len(tail_lineage) == 8:
                    tail_lineage = tail_lineage[:-1]
            if (
                head_lineage in taxons_in_otu_table
                and tail_lineage in taxons_in_otu_table
            ):
                interactions[tuple(sorted([head_lineage, tail_lineage]))].add(
                    interaction
                )
    return interactions


def read_predicted_interactions(interactions_fpath: str) -> dict[str, dict[str, str]]:
    interactions = dict()
    with open(interactions_fpath) as handle:
        reader = csv.reader(handle, delimiter="\t")
        next(reader)
        for head, tail, _, weight, sign in reader:
            interactions[tuple(sorted([head, tail]))] = {"sign": sign, "weight": weight}
    return interactions


def read_taxonomy(tax_fpath: str) -> dict[str, tuple[str, ...]]:
    taxonomy = dict()
    with open(tax_fpath) as handle:
        reader = csv.reader(handle, delimiter="\t")
        next(reader)
        for otu_id, *lineage in reader:
            taxonomy[otu_id] = tuple(lineage)
    return taxonomy


def get_prop_known_interactions(
    graph: nx.Graph,
    known_interactions: Dict[Tuple[Taxonomy, Taxonomy], Set[str]],
    taxonomy: dict[str, Taxonomy],
    count: bool = False,
    prop_of_network_edges: bool = False,
    kind="all",
    trim_to_genus: bool = False,
) -> float:
    n = 0
    results = Counter()
    unique_results = defaultdict(set)
    all_known_relations = sum(
        len(interaction_types) for interaction_types in known_interactions.values()
    )
    for head_lineage, tail_lineage in yield_tax_edges(graph, taxonomy, trim_to_genus):
        tax_relation = tuple(sorted([head_lineage, tail_lineage]))
        n += 1
        if tax_relation in known_interactions:
            data = known_interactions[tax_relation]  # noqa
            results["all"] += 1
            for interaction_type in data:
                unique_results["all"].add((tax_relation, interaction_type))
                results[interaction_type] += 1
                unique_results[interaction_type].add((tax_relation, interaction_type))
    if count:
        return len(unique_results[kind])
    else:
        if prop_of_network_edges:
            return results[kind] / n
        else:
            return len(unique_results[kind]) / all_known_relations


def yield_tax_edges(
    graph: Union[nx.Graph, dict[tuple[str, str], dict[str, str]]],
    taxonomy: dict[str, Taxonomy],
    trim_to_genus: bool,
    tax_in_OTU_table: Optional[set[Taxonomy]] = None,
) -> Iterator[Tuple[Taxonomy, Taxonomy]]:
    edge_iter = graph.edges() if isinstance(graph, nx.Graph) else iter(graph)
    for head, tail in edge_iter:
        if head in taxonomy and tail in taxonomy:
            head_tax = taxonomy[head]
            tail_tax = taxonomy[tail]
            if trim_to_genus:
                if len(head_tax) == 8:
                    head_tax = head_tax[:-1]
                if len(tail_tax) == 8:
                    tail_tax = tail_tax[:-1]
            if tax_in_OTU_table is None or (
                head_tax in tax_in_OTU_table and tail_tax in tax_in_OTU_table
            ):
                yield head_tax, tail_tax


def get_prop_predicted_interactions(
    graph: nx.Graph,
    predicted_interactions: dict[tuple[str, str], dict[str, str]],
    otu_ids: set[str],
    taxons_in_otu_table: set[Taxonomy],
    taxonomy: Optional[dict[str, Taxonomy]] = None,
    count: bool = False,
    prop_of_network_edges: bool = False,
    trim_to_genus: bool = False,
) -> float:
    if taxonomy is not None:
        inferred_interactions = Counter(
            tuple(sorted([head, tail]))
            for head, tail in yield_tax_edges(graph, taxonomy, trim_to_genus)
        )
        interactions = Counter(
            tuple(sorted([head, tail]))
            for head, tail in yield_tax_edges(
                predicted_interactions, taxonomy, trim_to_genus, taxons_in_otu_table
            )
        )
    else:
        inferred_interactions = Counter(
            tuple(sorted([head, tail])) for head, tail in graph.edges()
        )
        interactions = Counter(
            (head, tail)
            for head, tail in predicted_interactions
            if head in otu_ids and tail in otu_ids
        )
    common_interactions = inferred_interactions & interactions
    if count:
        return sum(common_interactions.values())
    else:
        if prop_of_network_edges:
            return sum(common_interactions.values()) / sum(
                inferred_interactions.values()
            )
        else:
            return sum(common_interactions.values()) / sum(interactions.values())


def get_n_signs(graph: nx.Graph, sign: str) -> int:
    n = 0
    for _, _, attrs in graph.edges(data=True):
        if attrs["sign"] == sign:
            n += 1
    return n


def return_if_except(exception, value: Any):
    def decorator(fun):
        def wrapper(*args, **kwargs):
            try:
                return fun(*args, **kwargs)
            except exception:
                return value

        return wrapper

    return decorator


@return_if_except(nx.NetworkXError, np.inf)
@return_if_except(ValueError, np.inf)
def get_radius(graph: nx.Graph) -> Union[int, float]:
    return nx.radius(graph)


@return_if_except(nx.NetworkXError, np.inf)
@return_if_except(ValueError, np.inf)
def get_diameter(graph: nx.Graph) -> Union[int, float]:
    return nx.diameter(graph)


@return_if_except(nx.NetworkXError, np.nan)
def get_avg_path_length(graph: nx.Graph) -> Union[int, float]:
    return nx.average_shortest_path_length(graph)


@return_if_except(ZeroDivisionError, np.nan)
def get_avg_clustering(graph: nx.Graph) -> Union[int, float]:
    return nx.average_clustering(graph)


def get_trophic_group_relations(
    graph: nx.Graph, trophic_groups: Dict[str, str], trophic_pair: Tuple[str, str]
) -> int:
    n = 0
    for head, tail in graph.edges():
        if head in trophic_groups and tail in trophic_groups:
            groups = [trophic_groups[head], trophic_groups[tail]]
            if tuple(sorted(groups)) == trophic_pair:
                n += 1
    return n


METRIC_TO_FUN: Dict[str, Callable[[nx.Graph], Union[float, int]]] = {
    "Number of nodes": lambda graph: graph.number_of_nodes(),
    "Number of edges": lambda graph: graph.number_of_edges(),
    "Number of positive interactions": partial(get_n_signs, sign="+"),
    "Number of negative interactions": partial(get_n_signs, sign="-"),
    "Number of connected components": lambda graph: nx.number_connected_components(
        graph
    ),
    "Mean network degree centrality": lambda graph: np.mean(
        list(nx.degree_centrality(graph).values())
    ),
    "Mean network betweeness": lambda graph: np.mean(
        list(nx.betweenness_centrality(graph).values())
    ),
    "Network radius": get_radius,
    "Network diameter": get_diameter,
    "Average clustering coefficient": get_avg_clustering,
    "Average path length": get_avg_path_length,
}

HEADER = list(METRIC_TO_FUN.keys())


def extend_metrics(
    known_interactions,
    known_interactions_genus,
    otu_ids: set[str],
    taxons_in_otu_table: set[Taxonomy],
    taxons_in_otu_table_genus: set[Taxonomy],
    predicted_interactions,
    taxonomy,
    trophic_groups,
):
    all_PIDA_interaction_types = set()
    for interaction_types in known_interactions.values():
        all_PIDA_interaction_types.update(interaction_types)
    for interaction_types in known_interactions_genus.values():
        all_PIDA_interaction_types.update(interaction_types)
    metrics = (
        [
            (
                "Discovered known interactions",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions,
                    taxonomy=taxonomy,
                    count=True,
                ),
            ),
            (
                "Proportion of known interactions discovered",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions,
                    taxonomy=taxonomy,
                ),
            ),
            (
                "Proportion of edges that appear as known interactions",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions,
                    taxonomy=taxonomy,
                    prop_of_network_edges=True,
                ),
            ),
        ]
        + [
            (
                f"Discovered known interactions ({kind})",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions,
                    taxonomy=taxonomy,
                    count=True,
                    kind=kind,
                ),
            )
            for kind in sorted(all_PIDA_interaction_types)
        ]
        + [
            (
                "Discovered known interactions (max genus level)",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions_genus,
                    taxonomy=taxonomy,
                    count=True,
                    trim_to_genus=True,
                ),
            ),
            (
                "Proportion of known interactions discovered (max genus level)",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions_genus,
                    taxonomy=taxonomy,
                    trim_to_genus=True,
                ),
            ),
            (
                "Proportion of edges that appear as known interactions (max genus level)",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions_genus,
                    taxonomy=taxonomy,
                    prop_of_network_edges=True,
                    trim_to_genus=True,
                ),
            ),
        ]
        + [
            (
                f"Discovered known interactions ({kind}) (max genus level)",
                partial(
                    get_prop_known_interactions,
                    known_interactions=known_interactions_genus,
                    taxonomy=taxonomy,
                    count=True,
                    kind=kind,
                    trim_to_genus=True,
                ),
            )
            for kind in sorted(all_PIDA_interaction_types)
        ]
        + [
            (
                "Discovered Lima-Mendez interactions (OTU level)",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table,
                    taxonomy=None,
                    count=True,
                ),
            ),
            (
                "Proportion of Lima-Mendez interactions (OTU level) discovered",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=None,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table,
                ),
            ),
            (
                "Proportion of edges that appear as Lima-Mendez interactions (OTU level)",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=None,
                    prop_of_network_edges=True,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table,
                ),
            ),
            (
                "Discovered Lima-Mendez interactions (taxonomy level)",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=taxonomy,
                    count=True,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table,
                ),
            ),
            (
                "Proportion of Lima-Mendez interactions (taxonomy level) discovered",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=taxonomy,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table,
                ),
            ),
            (
                "Proportion of edges that appear as Lima-Mendez interactions (taxonomy level)",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=taxonomy,
                    prop_of_network_edges=True,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table,
                ),
            ),
            (
                "Discovered Lima-Mendez interactions (max genus level)",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=taxonomy,
                    count=True,
                    trim_to_genus=True,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table_genus,
                ),
            ),
            (
                "Proportion of Lima-Mendez interactions (max genus level) discovered",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=taxonomy,
                    trim_to_genus=True,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table_genus,
                ),
            ),
            (
                "Proportion of edges that appear as Lima-Mendez interactions (max genus level)",
                partial(
                    get_prop_predicted_interactions,
                    predicted_interactions=predicted_interactions,
                    taxonomy=taxonomy,
                    prop_of_network_edges=True,
                    trim_to_genus=True,
                    otu_ids=otu_ids,
                    taxons_in_otu_table=taxons_in_otu_table_genus,
                ),
            ),
        ]
        + [
            (
                f"{group1}-{group2} interactions",
                partial(
                    get_trophic_group_relations,
                    trophic_groups=trophic_groups,
                    trophic_pair=tuple(sorted([group1, group2])),
                ),
            )
            for group1, group2 in combinations(sorted(set(trophic_groups.values())), 2)
        ]
    )
    for metric_name, fun in metrics:
        METRIC_TO_FUN[metric_name] = fun
        HEADER.append(metric_name)


METHOD_CAPITALIZATION = {
    "fastspar_results": "FastSpar (SparCC)",
    "flashweave_results": "FlashWeave",
    "phyloseq_results": "phyloseq",
    "conet_results": "Custom CoNet",
}


def prepare_stats(graph: nx.Graph):
    return {name: fun(graph) for name, fun in METRIC_TO_FUN.items()}


def read_graph(fpath: str) -> nx.Graph:
    graph = nx.Graph()
    with open(fpath) as handle:
        reader = csv.reader(handle, delimiter="\t")
        for node1, node2, *attrs in reader:
            if node1 != "leftover_vector" and node2 != "leftover_vector":
                graph.add_edge(node1, node2, sign=attrs[1], weight=attrs[0])
    return graph


def process(fpath: str):
    graph = read_graph(fpath)
    return fpath, prepare_stats(graph), graph


def write_interactions(
    graph: nx.Graph,
    graph_name: str,
    handle: OpenPyXLWriter,
    taxonomy: Dict[str, Tuple[str, ...]],
    known_interactions,
    known_interactions_genus,
    predicted_interactions,
    predicted_interactions_tax,
    predicted_interactions_genus,
    trophic_groups,
):
    handle.writerow(
        [
            "otu 1",
            "otu 2",
            "weight",
            "sign",
            "in PIDA",
            "in PIDA (genus level)",
            "PIDA interactions",
            "PIDA interactions (genus level)",
            "in Lima-Mendez",
            "Lima-Mendez sign",
            "in Lima-Mendez (taxon level)",
            "in Lima-Mendez (genus level)",
            "trophic group 1",
            "trophic group 2",
        ]
        + list(chain(*[[f"{rank} {i}" for rank in RANKS] for i in range(1, 3)])),
        graph_name,
    )
    for head, tail, attrs in graph.edges(data=True):
        head, tail = sorted([head, tail])
        head_taxonomy = taxonomy[head]
        tail_taxonomy = taxonomy[tail]
        head_taxonomy_genus = (
            head_taxonomy[:-1] if len(head_taxonomy) == 8 else head_taxonomy
        )
        tail_taxonomy_genus = (
            tail_taxonomy[:-1] if len(tail_taxonomy) == 8 else tail_taxonomy
        )
        key = tuple(sorted([head_taxonomy, tail_taxonomy]))
        key_genus = tuple(sorted([head_taxonomy_genus, tail_taxonomy_genus]))
        if key in known_interactions:
            known = True
            interaction = ", ".join(known_interactions[key])
        else:
            known = False
            interaction = ""
        if key_genus in known_interactions_genus:
            known_genus = True
            interaction_genus = ", ".join(known_interactions_genus[key_genus])
        else:
            known_genus = False
            interaction_genus = ""
        if (head, tail) in predicted_interactions:
            predicted = True
            predicted_sign = predicted_interactions[(head, tail)]["sign"]
        else:
            predicted = False
            predicted_sign = ""
        if key in predicted_interactions_tax:
            predicted_tax = True
        else:
            predicted_tax = False
        if key_genus in predicted_interactions_genus:
            predicted_genus = True
        else:
            predicted_genus = False
        row = [
            head,
            tail,
            attrs["weight"],
            attrs["sign"],
            known,
            interaction,
            known_genus,
            interaction_genus,
            predicted,
            predicted_sign,
            predicted_tax,
            predicted_genus,
            trophic_groups.get(head, ""),
            trophic_groups.get(tail, ""),
            *(
                list(head_taxonomy)
                + ["" for _ in range(len(RANKS) - len(head_taxonomy))]
            ),
            *(
                list(tail_taxonomy)
                + ["" for _ in range(len(RANKS) - len(tail_taxonomy))]
            ),
        ]
        handle.writerow(row, graph_name)


if __name__ == "__main__":
    (
        otu_table_fpath,
        tax_path,
        known_inter_path,
        known_inter_path_genus,
        pred_path,
        trophic_groups_path,
        n_threads,
        output_fpath,
        *network_fpaths,
    ) = sys.argv[1:]
    proc_delayed = delayed(process)
    executor = Parallel(n_jobs=int(n_threads), verbose=5)

    taxonomy = read_taxonomy(tax_path)
    taxons_in_otu_table = set()
    taxons_in_otu_table_genus = set()
    otu_ids = set()
    with open(otu_table_fpath) as handle:
        reader = csv.reader(handle, delimiter="\t")
        next(reader)
        for otu_id, *_ in reader:
            if otu_id != "leftover_vector":
                otu_ids.add(otu_id)
                lineage = taxonomy[otu_id]
                taxons_in_otu_table.add(lineage)
                if len(lineage) == 8:
                    taxons_in_otu_table_genus.add(lineage[:-1])
                else:
                    taxons_in_otu_table_genus.add(lineage)
    known_interactions = read_experimental_interactions(
        known_inter_path, taxons_in_otu_table
    )
    known_interactions_genus = read_experimental_interactions(
        known_inter_path_genus, taxons_in_otu_table_genus, True
    )
    predicted_interactions = read_predicted_interactions(pred_path)
    predicted_interactions_tax = set()
    predicted_interactions_genus = set()
    for head, tail in predicted_interactions:
        head_tax = taxonomy[head]
        tail_tax = taxonomy[tail]
        predicted_interactions_tax.add(tuple(sorted([head_tax, tail_tax])))
        if len(head_tax) == 8:
            head_tax = head_tax[:-1]
        if len(tail_tax) == 8:
            tail_tax = tail_tax[:-1]
        predicted_interactions_genus.add(tuple(sorted([head_tax, tail_tax])))
    trophic_groups = read_trophic_groups(trophic_groups_path, taxonomy)
    extend_metrics(
        known_interactions,
        known_interactions_genus,
        otu_ids,
        taxons_in_otu_table,
        taxons_in_otu_table_genus,
        predicted_interactions,
        taxonomy,
        trophic_groups,
    )
    tasks = (proc_delayed(fpath) for fpath in network_fpaths)
    graphs_by_method = {}
    stats_by_method = {}
    for fpath, stats, graph in executor(tasks):
        if "consensus_network" in fpath:
            network_name = "Consensus network"
        else:
            head, tail = os.path.split(fpath)
            method_name = METHOD_CAPITALIZATION[os.path.split(head)[1]]
            network_name = method_name + " network"
        graphs_by_method[network_name] = graph
        stats_by_method[network_name] = stats
    lima_mendez_graph = nx.Graph()
    for (head, tail), attrs in predicted_interactions.items():
        lima_mendez_graph.add_edge(head, tail, **attrs)
    stats_by_method["Lima-Mendez TARA interactome"] = prepare_stats(lima_mendez_graph)
    graphs_by_method["Lima-Mendez TARA interactome"] = lima_mendez_graph
    graphs = []
    names = []
    for graph_name, graph in graphs_by_method.items():
        graphs.append(graph)
        names.append(graph_name)
    hub_stats = prepare_hub_taxons_stats(graphs, names, taxonomy)

    with OpenPyXLWriter(output_fpath, "Network statistics") as handle:
        handle.writerow(["network"] + HEADER)
        for network_name, stats in stats_by_method.items():
            handle.writerow(
                [network_name] + [stats[statistic] for statistic in HEADER],
                "Network statistics",
            )
        prepare_pairwise_stats(graphs_by_method, handle, taxonomy)
        for network_name, graph in graphs_by_method.items():
            write_interactions(
                graph,
                network_name,
                handle,
                taxonomy,
                known_interactions,
                known_interactions_genus,
                predicted_interactions,
                predicted_interactions_tax,
                predicted_interactions_genus,
                trophic_groups,
            )
        write_hub_stats(handle, names, hub_stats)
